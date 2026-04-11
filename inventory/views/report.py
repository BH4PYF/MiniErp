from datetime import date
from decimal import Decimal
import logging

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.db.models import Sum
from django.db.models.functions import ExtractYear, ExtractMonth
from django.views.decorators.http import require_GET
from django.core.cache import cache

logger = logging.getLogger(__name__)


from ..models import (
    Project, Category, Supplier,
    InboundRecord, Subcontractor, Contract, Measurement, Settlement,
)
from .utils import (
    log_operation, parse_date, role_required,
    create_excel_workbook, make_excel_response,
)


@role_required('admin', 'management')
def report_page(request):
    projects = Project.objects.all()
    categories = Category.objects.all().order_by('code')
    suppliers = Supplier.objects.all()
    subcontractors = Subcontractor.objects.all()
    return render(request, 'inventory/report.html', {
        'projects': projects, 'categories': categories, 'suppliers': suppliers,
        'subcontractors': subcontractors,
    })


@role_required('admin', 'management')
def report_project_cost(request):
    project_id = request.GET.get('project_id')
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))
    export = request.GET.get('export', '')

    if not project_id:
        return JsonResponse({'error': '请选择项目'}, status=400)

    project = get_object_or_404(Project, pk=project_id)

    # 构建缓存键
    cache_key = f'report_project_cost_{project_id}_{date_from or ""}_{date_to or ""}'
    cached_data = cache.get(cache_key)

    if cached_data and export != 'excel':
        material_agg = cached_data['material_agg']
        total = cached_data['total']
    else:
        # 使用数据库层聚合代替 Python 层聚合
        inbounds = InboundRecord.objects.filter(project=project).select_related('material', 'material__category')
        if date_from:
            inbounds = inbounds.filter(date__gte=date_from)
        if date_to:
            inbounds = inbounds.filter(date__lte=date_to)

        # 在数据库层面按材料分组聚合，避免加载所有记录到内存
        material_agg = inbounds.values(
            'material_id',
            'material__name',
            'material__spec',
            'material__unit',
            'material__category__name'
        ).annotate(
            total_qty=Sum('quantity'),
            total_cost=Sum('total_amount')
        ).order_by('-total_cost')
        
        # 计算总成本
        total = sum(row['total_cost'] for row in material_agg) or Decimal('0')
        
        # 缓存数据，有效期10分钟
        if export != 'excel':
            cache.set(cache_key, {'material_agg': material_agg, 'total': total}, 600)

    if export == 'excel':
        headers = ['分类', '材料名称', '规格', '单位', '数量', '平均单价', '金额', '占比']
        wb, ws, border = create_excel_workbook(f'{project.name}_采购成本分析', headers, style='report')
        row = 2
        for v in material_agg:
            ratio = f"{v['total_cost'] / total * 100:.1f}%" if total else '0%'
            # 计算平均单价
            avg_price = v['total_cost'] / v['total_qty'] if v['total_qty'] > 0 else 0
            ws.cell(row=row, column=1, value=v['material__category__name']).border = border
            ws.cell(row=row, column=2, value=v['material__name']).border = border
            ws.cell(row=row, column=3, value=v['material__spec']).border = border
            ws.cell(row=row, column=4, value=v['material__unit']).border = border
            ws.cell(row=row, column=5, value=float(v['total_qty'])).border = border
            ws.cell(row=row, column=6, value=float(avg_price)).border = border
            ws.cell(row=row, column=7, value=float(v['total_cost'])).border = border
            ws.cell(row=row, column=8, value=ratio).border = border
            row += 1
        from openpyxl.styles import Font
        ws.cell(row=row, column=5, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=7, value=float(total)).font = Font(bold=True)
        filename = f'{project.name}_采购成本分析_{date_from}_{date_to}.xlsx'
        log_operation(request.user, '统计报表', 'export', f'导出项目采购成本分析 {project.name} 期间:{date_from}至{date_to}', project.code)
        return make_excel_response(wb, filename)
    else:
        # 转换为模板需要的格式
        report_data = []
        for v in material_agg:
            # 计算平均单价
            avg_price = v['total_cost'] / v['total_qty'] if v['total_qty'] > 0 else Decimal('0')
            report_data.append({
                'category': v['material__category__name'],
                'name': v['material__name'],
                'spec': v['material__spec'],
                'unit': v['material__unit'],
                'quantity': v['total_qty'],
                'avg_price': avg_price,
                'cost': v['total_cost'],
                'ratio': v['total_cost'] / total * 100 if total else 0,
            })
        return render(request, 'inventory/report_project_cost.html', {
            'project': project, 'report_data': report_data, 'total_cost': total,
            'date_from': date_from, 'date_to': date_to,
            'projects': Project.objects.all(),
        })


@role_required('admin', 'management')
def report_supplier_cost(request):
    supplier_id = request.GET.get('supplier_id')
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))
    export = request.GET.get('export', '')

    if not supplier_id:
        return JsonResponse({'error': '请选择供应商'}, status=400)

    supplier = get_object_or_404(Supplier, pk=supplier_id)

    # 构建缓存键
    cache_key = f'report_supplier_cost_{supplier_id}_{date_from or ""}_{date_to or ""}'
    cached_data = cache.get(cache_key)

    if cached_data and export != 'excel':
        inbounds = cached_data['inbounds']
        total_amount = cached_data['total_amount']
    else:
        inbounds = InboundRecord.objects.filter(supplier=supplier).select_related('material', 'project')
        if date_from:
            inbounds = inbounds.filter(date__gte=date_from)
        if date_to:
            inbounds = inbounds.filter(date__lte=date_to)

        total_amount = inbounds.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # 缓存数据，有效期10分钟
        if export != 'excel':
            cache.set(cache_key, {'inbounds': inbounds, 'total_amount': total_amount}, 600)

    if export == 'excel':
        MAX_EXPORT_ROWS = 10000
        headers = ['项目', '材料名称', '规格', '单位', '数量', '单价', '总金额', '日期']
        wb, ws, border = create_excel_workbook(f'{supplier.name}采购分析', headers, style='report')
        row = 2
        total_amount = Decimal('0')
        for r in inbounds.iterator(chunk_size=500):
            if row - 2 >= MAX_EXPORT_ROWS:
                break
            ws.cell(row=row, column=1, value=r.project.name).border = border
            ws.cell(row=row, column=2, value=r.material.name).border = border
            ws.cell(row=row, column=3, value=r.material.spec).border = border
            ws.cell(row=row, column=4, value=r.material.unit).border = border
            ws.cell(row=row, column=5, value=float(r.quantity)).border = border
            ws.cell(row=row, column=6, value=float(r.unit_price)).border = border
            ws.cell(row=row, column=7, value=float(r.total_amount)).border = border
            ws.cell(row=row, column=8, value=r.date.strftime('%Y-%m-%d')).border = border
            total_amount += r.total_amount
            row += 1
        from openpyxl.styles import Font
        ws.cell(row=row, column=6, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=7, value=float(total_amount)).font = Font(bold=True)
        filename = f'{supplier.name}_采购分析_{date_from}_{date_to}.xlsx'
        log_operation(request.user, '统计报表', 'export', f'导出供应商采购分析 {supplier.name} 期间:{date_from}至{date_to}', supplier.code)
        return make_excel_response(wb, filename)
    else:
        from django.core.paginator import Paginator

        paginator = Paginator(inbounds, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        records = []
        for r in page_obj:
            records.append({
                'project': r.project.name,
                'material': r.material.name,
                'spec': r.material.spec,
                'unit': r.material.unit,
                'quantity': r.quantity,
                'unit_price': r.unit_price,
                'total_amount': r.total_amount,
                'date': r.date,
            })

        return render(request, 'inventory/report_supplier_cost.html', {
            'supplier': supplier, 'records': records, 'total_amount': total_amount,
            'date_from': date_from, 'date_to': date_to,
            'suppliers': Supplier.objects.all(),
            'page_obj': page_obj,
        })


@role_required('admin', 'management')
def report_monthly(request):
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))
    export = request.GET.get('export', '')

    # 构建缓存键
    cache_key = f'report_monthly_{date_from or ""}_{date_to or ""}'
    cached_data = cache.get(cache_key)

    if cached_data and export != 'excel':
        date_data = cached_data['date_data']
    else:
        qs = InboundRecord.objects.all()
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # 在数据库层面按日期聚合，避免全量加载到内存
        date_agg = qs.values('date').annotate(
            total=Sum('total_amount')
        ).order_by('date')
        date_data = {row['date']: row['total'] for row in date_agg}
        
        # 缓存数据，有效期10分钟
        if export != 'excel':
            cache.set(cache_key, {'date_data': date_data}, 600)

    if export == 'excel':
        headers = ['日期', '入库金额']
        wb, ws, border = create_excel_workbook(f'入库统计_{date_from}_{date_to}', headers, style='report')

        row = 2
        total_in = Decimal('0')
        sorted_dates = sorted(date_data.keys())
        for d in sorted_dates:
            ws.cell(row=row, column=1, value=d.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=float(date_data[d])).border = border
            total_in += date_data[d]
            row += 1
        from openpyxl.styles import Font
        ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(total_in)).font = Font(bold=True)
        filename = f'入库统计_{date_from}_{date_to}.xlsx'
        log_operation(request.user, '统计报表', 'export', f'导出入库统计 期间:{date_from}至{date_to}')
        return make_excel_response(wb, filename)
    else:
        daily_data = []
        total_in = Decimal('0')
        sorted_dates = sorted(date_data.keys())
        for d in sorted_dates:
            daily_data.append({
                'date': d,
                'inbound': date_data[d],
            })
            total_in += date_data[d]

        return render(request, 'inventory/report_monthly.html', {
            'report_data': daily_data, 'total_in': total_in,
            'date_from': date_from, 'date_to': date_to,
        })


# ========== 图表分析 ==========

@role_required('admin', 'management')
def chart_page(request):
    return render(request, 'inventory/charts.html')


@role_required('admin', 'management')
@require_GET
def chart_data_api(request):
    chart_type = request.GET.get('type', 'stock')
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))

    if date_from and date_to:
        start_date = date_from
        end_date = date_to
    else:
        start_date = None
        end_date = None

    if chart_type == 'stock':
        qs = InboundRecord.objects.all()
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])
        agg = qs.values('material__name').annotate(
            total_amount_sum=Sum('total_amount'),
        ).filter(total_amount_sum__gt=0).order_by('-total_amount_sum')[:10]
        data = [{'name': row['material__name'], 'value': float(row['total_amount_sum'])} for row in agg]
        return JsonResponse({'data': data})
    elif chart_type == 'category':
        qs = InboundRecord.objects.all()
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])
        agg = qs.values('material__category__name').annotate(
            total_value=Sum('total_amount'),
        ).filter(total_value__gt=0)
        data = [{'name': row['material__category__name'], 'value': float(row['total_value'])} for row in agg]
        return JsonResponse({'data': data})
    elif chart_type == 'inbound_monthly':
        year_raw = request.GET.get('year', '')
        try:
            year = int(year_raw) if year_raw else date.today().year
        except ValueError:
            return JsonResponse({'error': '年份参数格式错误'}, status=400)

        qs = InboundRecord.objects.filter(date__year=year)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        month_agg = qs.annotate(month=ExtractMonth('date')).values('month').annotate(
            total=Sum('total_amount')
        )
        month_map = {row['month']: row['total'] for row in month_agg}
        data = [{'month': i, 'value': float(month_map.get(i, 0) or 0)} for i in range(1, 13)]
        return JsonResponse({'data': data})
    return JsonResponse({'data': []})


@role_required('admin', 'management')
@require_GET
def get_years_list(request):
    """返回有入库数据的年份列表（降序）"""
    years = list(
        InboundRecord.objects.annotate(y=ExtractYear('date'))
        .values_list('y', flat=True)
        .distinct()
        .order_by('-y')
    )
    if not years:
        years = [date.today().year]
    return JsonResponse({'years': years})


# ========== 分包分析报表 ==========

@role_required('admin', 'management')
def report_project_subcontract(request):
    """项目分包分析"""
    project_id = request.GET.get('project_id')
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))
    export = request.GET.get('export', '')

    if not project_id:
        return JsonResponse({'error': '请选择项目'}, status=400)

    project = get_object_or_404(Project, pk=project_id)

    # 构建缓存键
    cache_key = f'report_project_subcontract_{project_id}_{date_from or ""}_{date_to or ""}'
    cached_data = cache.get(cache_key)

    if cached_data and export != 'excel':
        report_data = cached_data['report_data']
        contract_total = cached_data['contract_total']
        settlement_total = cached_data['settlement_total']
    else:
        # 统计项目的分包合同和结算
        contracts = Contract.objects.filter(project=project).select_related('subcontractor')
        settlements = Settlement.objects.filter(contract__project=project).select_related('contract', 'contract__subcontractor')

        if date_from:
            contracts = contracts.filter(created_at__date__gte=date_from)
            settlements = settlements.filter(period_end__gte=date_from)
        if date_to:
            contracts = contracts.filter(created_at__date__lte=date_to)
            settlements = settlements.filter(period_end__lte=date_to)

        # 计算合同总额
        contract_total = Decimal('0')
        try:
            for contract in contracts:
                try:
                    contract_total += contract.get_contract_total()
                except (ValueError, AttributeError) as e:
                    logger.warning(f"合同 {contract.id} 金额计算错误: {e}")
        except Exception as e:
            logger.error(f"合同总额计算失败: {e}")
        # 计算结算总额
        settlement_total = Decimal('0')
        try:
            settlement_total = settlements.aggregate(total=Sum('final_amount'))['total'] or Decimal('0')
        except (ValueError, TypeError) as e:
            logger.error(f"结算总额计算错误: {e}")
            settlement_total = Decimal('0')

        # 按分包商分组统计
        subcontractor_stats = {}
        for contract in contracts:
            sub_name = contract.subcontractor.name
            if sub_name not in subcontractor_stats:
                subcontractor_stats[sub_name] = {'contract': Decimal('0'), 'settlement': Decimal('0')}
            subcontractor_stats[sub_name]['contract'] += contract.get_contract_total()

        for settlement in settlements:
            sub_name = settlement.contract.subcontractor.name
            if sub_name not in subcontractor_stats:
                subcontractor_stats[sub_name] = {'contract': Decimal('0'), 'settlement': Decimal('0')}
            subcontractor_stats[sub_name]['settlement'] += settlement.final_amount or Decimal('0')

        # 转换为列表格式
        report_data = []
        for sub_name, stats in subcontractor_stats.items():
            report_data.append({
                'subcontractor': sub_name,
                'contract_amount': stats['contract'],
                'settlement_amount': stats['settlement'],
                'balance': stats['contract'] - stats['settlement'],
            })

        # 按合同金额排序
        report_data.sort(key=lambda x: x['contract_amount'], reverse=True)
        
        # 缓存数据，有效期10分钟
        if export != 'excel':
            cache.set(cache_key, {
                'report_data': report_data,
                'contract_total': contract_total,
                'settlement_total': settlement_total
            }, 600)

    if export == 'excel':
        headers = ['分包商', '合同金额', '结算金额', '未结算金额']
        wb, ws, border = create_excel_workbook(f'{project.name}_分包分析', headers, style='report')
        row = 2
        for item in report_data:
            ws.cell(row=row, column=1, value=item['subcontractor']).border = border
            ws.cell(row=row, column=2, value=float(item['contract_amount'])).border = border
            ws.cell(row=row, column=3, value=float(item['settlement_amount'])).border = border
            ws.cell(row=row, column=4, value=float(item['balance'])).border = border
            row += 1
        from openpyxl.styles import Font
        ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(contract_total)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(settlement_total)).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(contract_total - settlement_total)).font = Font(bold=True)
        filename = f'{project.name}_分包分析_{date_from}_{date_to}.xlsx'
        log_operation(request.user, '统计报表', 'export', f'导出项目分包分析 {project.name} 期间:{date_from}至{date_to}', project.code)
        return make_excel_response(wb, filename)
    else:
        return render(request, 'inventory/report_project_subcontract.html', {
            'project': project, 'report_data': report_data,
            'contract_total': contract_total, 'settlement_total': settlement_total,
            'balance_total': contract_total - settlement_total,
            'date_from': date_from, 'date_to': date_to,
            'projects': Project.objects.all(),
        })


@role_required('admin', 'management')
def report_subcontractor_analysis(request):
    """分包商分包分析"""
    subcontractor_id = request.GET.get('subcontractor_id')
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))
    export = request.GET.get('export', '')

    if not subcontractor_id:
        return JsonResponse({'error': '请选择分包商'}, status=400)

    subcontractor = get_object_or_404(Subcontractor, pk=subcontractor_id)

    # 构建缓存键
    cache_key = f'report_subcontractor_analysis_{subcontractor_id}_{date_from or ""}_{date_to or ""}'
    cached_data = cache.get(cache_key)

    if cached_data and export != 'excel':
        report_data = cached_data['report_data']
        contract_total = cached_data['contract_total']
        settlement_total = cached_data['settlement_total']
        balance_total = cached_data['balance_total']
    else:
        # 统计分包商的合同和结算
        contracts = Contract.objects.filter(subcontractor=subcontractor).select_related('project')
        settlements = Settlement.objects.filter(contract__subcontractor=subcontractor).select_related('contract', 'contract__project')

        if date_from:
            contracts = contracts.filter(created_at__date__gte=date_from)
            settlements = settlements.filter(period_end__gte=date_from)
        if date_to:
            contracts = contracts.filter(created_at__date__lte=date_to)
            settlements = settlements.filter(period_end__lte=date_to)

        # 按项目分组统计
        project_stats = {}
        try:
            for contract in contracts:
                try:
                    project_name = f"{contract.project.code} - {contract.project.name}"
                    if project_name not in project_stats:
                        project_stats[project_name] = {'contract': Decimal('0'), 'settlement': Decimal('0')}
                    project_stats[project_name]['contract'] += contract.get_contract_total()
                except (ValueError, AttributeError) as e:
                    logger.warning(f"合同 {contract.id} 项目分组统计错误: {e}")

            for settlement in settlements:
                try:
                    project_name = f"{settlement.contract.project.code} - {settlement.contract.project.name}"
                    if project_name not in project_stats:
                        project_stats[project_name] = {'contract': Decimal('0'), 'settlement': Decimal('0')}
                    project_stats[project_name]['settlement'] += settlement.final_amount or Decimal('0')
                except (ValueError, AttributeError) as e:
                    logger.warning(f"结算 {settlement.id} 项目分组统计错误: {e}")
        except Exception as e:
            logger.error(f"项目分组统计失败: {e}")

        # 转换为列表格式
        report_data = []
        for project_name, stats in project_stats.items():
            report_data.append({
                'project': project_name,
                'contract_amount': stats['contract'],
                'settlement_amount': stats['settlement'],
                'balance': stats['contract'] - stats['settlement'],
            })

        # 按合同金额排序
        report_data.sort(key=lambda x: x['contract_amount'], reverse=True)

        # 计算总计
        contract_total = sum(item['contract_amount'] for item in report_data)
        settlement_total = sum(item['settlement_amount'] for item in report_data)
        balance_total = contract_total - settlement_total
        
        # 缓存数据，有效期10分钟
        if export != 'excel':
            cache.set(cache_key, {
                'report_data': report_data,
                'contract_total': contract_total,
                'settlement_total': settlement_total,
                'balance_total': balance_total
            }, 600)

    if export == 'excel':
        headers = ['项目', '合同金额', '结算金额', '未结算金额']
        wb, ws, border = create_excel_workbook(f'{subcontractor.name}_分包分析', headers, style='report')
        row = 2
        for item in report_data:
            ws.cell(row=row, column=1, value=item['project']).border = border
            ws.cell(row=row, column=2, value=float(item['contract_amount'])).border = border
            ws.cell(row=row, column=3, value=float(item['settlement_amount'])).border = border
            ws.cell(row=row, column=4, value=float(item['balance'])).border = border
            row += 1
        from openpyxl.styles import Font
        ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(contract_total)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(settlement_total)).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(balance_total)).font = Font(bold=True)
        filename = f'{subcontractor.name}_分包分析_{date_from}_{date_to}.xlsx'
        log_operation(request.user, '统计报表', 'export', f'导出分包商分包分析 {subcontractor.name} 期间:{date_from}至{date_to}', subcontractor.code)
        return make_excel_response(wb, filename)
    else:
        return render(request, 'inventory/report_subcontractor_analysis.html', {
            'subcontractor': subcontractor, 'report_data': report_data,
            'contract_total': contract_total, 'settlement_total': settlement_total,
            'balance_total': balance_total,
            'date_from': date_from, 'date_to': date_to,
            'subcontractors': Subcontractor.objects.all(),
        })


@role_required('admin', 'management')
def report_monthly_subcontract(request):
    """月度分包查询"""
    date_from = parse_date(request.GET.get('date_from', ''))
    date_to = parse_date(request.GET.get('date_to', ''))
    export = request.GET.get('export', '')

    # 构建缓存键
    cache_key = f'report_monthly_subcontract_{date_from or ""}_{date_to or ""}'
    cached_data = cache.get(cache_key)

    if cached_data and export != 'excel':
        report_data = cached_data['report_data']
        total_amount = cached_data['total_amount']
        total_count = cached_data['total_count']
    else:
        # 统计月度分包数据
        settlements = Settlement.objects.select_related('contract', 'contract__project', 'contract__subcontractor')
        if date_from:
            settlements = settlements.filter(period_end__gte=date_from)
        if date_to:
            settlements = settlements.filter(period_end__lte=date_to)

        # 按月份分组
        monthly_data = {}
        for settlement in settlements:
            month_key = settlement.period_end.strftime('%Y-%m')
            if month_key not in monthly_data:
                monthly_data[month_key] = {'amount': Decimal('0'), 'count': 0}
            monthly_data[month_key]['amount'] += settlement.final_amount or Decimal('0')
            monthly_data[month_key]['count'] += 1

        # 转换为列表并排序
        report_data = []
        for month, data in sorted(monthly_data.items()):
            report_data.append({
                'month': month,
                'amount': data['amount'],
                'count': data['count'],
            })

        # 计算总计
        total_amount = sum(item['amount'] for item in report_data)
        total_count = sum(item['count'] for item in report_data)
        
        # 缓存数据，有效期10分钟
        if export != 'excel':
            cache.set(cache_key, {
                'report_data': report_data,
                'total_amount': total_amount,
                'total_count': total_count
            }, 600)

    if export == 'excel':
        headers = ['月份', '结算金额', '结算次数']
        wb, ws, border = create_excel_workbook(f'月度分包统计_{date_from}_{date_to}', headers, style='report')
        row = 2
        for item in report_data:
            ws.cell(row=row, column=1, value=item['month']).border = border
            ws.cell(row=row, column=2, value=float(item['amount'])).border = border
            ws.cell(row=row, column=3, value=item['count']).border = border
            row += 1
        from openpyxl.styles import Font
        ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(total_amount)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_count).font = Font(bold=True)
        filename = f'月度分包统计_{date_from}_{date_to}.xlsx'
        log_operation(request.user, '统计报表', 'export', f'导出月度分包统计 期间:{date_from}至{date_to}')
        return make_excel_response(wb, filename)
    else:
        return render(request, 'inventory/report_monthly_subcontract.html', {
            'report_data': report_data, 'total_amount': total_amount, 'total_count': total_count,
            'date_from': date_from, 'date_to': date_to,
        })
